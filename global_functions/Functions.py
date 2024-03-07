import json
import os
import time
from datetime import datetime
from typing import Dict, Any

import openpyxl
import pytest
from appium.options.common import AppiumOptions
from appium.webdriver.appium_service import AppiumService
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException, MoveTargetOutOfBoundsException
from selenium.webdriver.common import service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from webdriver_manager.opera import OperaDriverManager
from selenium.webdriver.chrome import service
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver import ChromeOptions as OpcionesChrome, ActionChains
from Config.Config import Config
from Config.Config_execution import *


class Functions(Config):
    ##########################################################################
    ##############   -=_Config DRIVERS_=-   #############################
    ##########################################################################
    def start_appium(self):
        """
        Whith this step you can start the execution the appium server 
        """
        self.appium_service = AppiumService()
        try:
            self.appium_service.stop()
        except:
            pass
        self.appium_service.start(args=["--base-path","/wd/hub"])
        print("Starting appium service")

    def stop_appium(self):
        """
        Whith this step you can stop the execution the appium server 
        """
        self.appium_service.stop()
        print("Stopping appium service")

    def create_appium_driver(self, URL=Config.URL, navegador=Config.browser):
        """
        Whith this step you can create a new driver for the execution
        
        :param
            - URL: Site Web
            - navegador: Browser to do in the test 
        """
        print("\nDirectorio Base: " + Config.basedir)
        self.ventanas = {}
        print("----------------")
        print(navegador)
        print("---------------")

        cap_chrome: Dict[str, Any] = {
            "platformName": "Android",
            "automationName": "UiAutomator2",
            "browserName": "chrome",
            "unicodeKeyboard": True,
            "resetKeyboard": True,
            "noReset": True
        }
        cap_firefox: Dict[str, Any] = {
            "platformName": "windows",
            "automationName": "Gecko",
            "browserName": "Firefox",
            "marionette": True,
            "moz:firefoxOptions": {
                "androidPackage": "org.mozilla.firefox",
            }
        }
        cap_opera: Dict[str, Any] = {
            "platformName": "Android",
            "automationName": "UiAutomator2",
            "browserName": "chrome",
            "unicodeKeyboard": True,
            "resetKeyboard": True,
            "noReset": True,
            # "executablePath": Config.basedir + "\\drivers\\operadriver.exe",
            "chromeOptions": {
                # "binary": Config.basedir + "\\drivers\\operadriver.exe",
                # "androidPackage": "com.opera.browser",
                # "androidActivity": "com.opera.android.BrowserActivity",
            }
        }

        if navegador.lower() == "chrome":
            cap = cap_chrome
        elif navegador.lower() == "firefox":
            cap = cap_firefox
        elif navegador.lower() == "opera":
            cap = cap_opera

        self.driver = webdriver.Remote('http://127.0.0.1:4723/wd/hub', options=AppiumOptions().load_capabilities(cap))

        # if navegador.lower() == "chrome":
        #     options = OpcionesChrome()
        #     options.add_argument('start-maximized')
        #     self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        # self.driver.maximize_window()
        # elif navegador.lower() == "edge":
        #     self.driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))
        # self.driver.maximize_window()
        # elif navegador.lower() == "firefox":
        #     self.driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()))
        # self.driver.maximize_window()
        # elif navegador.lower() == "opera":
        #     webdriver_service = service.Service(OperaDriverManager().install())
        #     webdriver_service.start()
        #
        #     options = webdriver.ChromeOptions()
        #     options.add_experimental_option('w3c', True)
        #
        # self.driver = webdriver.Remote(webdriver_service.service_url, options=options)
        # self.driver.maximize_window()

        self.driver.implicitly_wait(10)
        self.driver.get(URL)
        self.ventanas = {'Principal': self.driver.window_handles[0]}
        return self.driver

    def tearDown(self):
        """
        Whith this step you can close the driver  
        """
        print("Close Driver")
        self.driver.quit()

    ##########################################################################
    ##############       -=_JSON     HANDLE _=-              #################
    ##########################################################################

    def get_json_file(self, file):
        """
        Whith this step you can get the content the json file of the elements 
        
        :param
            - file: name of the json file
        """
        json_path = Config.Json + "/" + file + '.json'
        try:
            with open(json_path, "r") as read_file:
                self.json_strings = json.loads(read_file.read())
        except FileNotFoundError:
            self.json_strings = False
            Functions.tearDown(self)

    def get_entity(self, entity):
        """
        Get the properties of the entity by create the element
        
        :param
            - entity: name of the element to retrieve properties from json file
        """
        if self.json_strings is False:
            print("Json not fond")
        else:
            try:
                self.json_ValueToFind = self.json_strings[entity]["ValueToFind"]
                self.json_GetFieldBy = self.json_strings[entity]["GetFieldBy"]
                return True

            except KeyError:
                Functions.tearDown(self)
                return None

    def get_elements(self, entity):
        """
        Return a page object by passing the element 
        
        :param
            - entity: name of the object to return
        
        return:
         -  page object
        """
        Get_Entity = Functions.get_entity(self, entity)

        if Get_Entity is None:
            print("Json isn't found")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    elements = self.driver.find_element(By.ID, self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "name":
                    elements = self.driver.find_element(By.NAME, self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "xpath":
                    elements = self.driver.find_element(By.XPATH, self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "link":
                    elements = self.driver.find_element(By.PARTIAL_LINK_TEXT, self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "css":
                    elements = self.driver.find_element(By.CSS_SELECTOR, self.json_ValueToFind)
                return elements

            except NoSuchElementException:
                print("get_text: Element not found: " + self.json_ValueToFind)
                Functions.tearDown(self)
            except TimeoutException:
                print("get_text: Element not found: " + self.json_ValueToFind)
                Functions.tearDown(self)

    def wait_element(self, locator):
        """
        Wait to element will be located in the front
        
        :param
            -locator: name of element to be located
        """
        Get_Entity = Functions.get_entity(self, locator)

        if Get_Entity is None:
            return print("Json not found")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    wait = WebDriverWait(self.driver, 20)
                    wait.until(EC.visibility_of_element_located((By.ID, self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.ID, self.json_ValueToFind)))
                    return True

                if self.json_GetFieldBy.lower() == "name":
                    wait = WebDriverWait(self.driver, 20)
                    wait.until(EC.visibility_of_element_located((By.NAME, self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.NAME, self.json_ValueToFind)))
                    return True

                if self.json_GetFieldBy.lower() == "xpath":
                    wait = WebDriverWait(self.driver, 20)

                    wait.until(EC.visibility_of_element_located((By.XPATH, self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.XPATH, self.json_ValueToFind)))
                    return True
                if self.json_GetFieldBy.lower() == "css":
                    wait = WebDriverWait(self.driver, 20)

                    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, self.json_ValueToFind)))
                    return True

                if self.json_GetFieldBy.lower() == "link":
                    wait = WebDriverWait(self.driver, 20)
                    wait.until(EC.visibility_of_element_located((By.PARTIAL_LINK_TEXT, self.json_ValueToFind)))
                    wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, self.json_ValueToFind)))
                    return True

            except TimeoutException:
                print("Wait Element: Element not found " + locator)
                Functions.tearDown(self)
            except NoSuchElementException:
                print("Wait Element: Element not foun" + locator)
                Functions.tearDown(self)

    def fill_excel_report(self, excel=Config.Excel):
        """
        this function uptades excel file with the result of the steps
        
        :param 
            excel: Excel name to updata  
        """
        value_col = []
        wb = openpyxl.load_workbook(Config.Excel)
        hoja = wb.active
        filas = hoja.max_row
        columnas = hoja.max_column
        Test_Case_ID = "validate " + self.texto
        Developer = "Developer" + str(filas)
        Test_group = "landing Page"
        Test_case_name = "Validate " + self.texto
        Priority_Severity = "1/4"
        Pre_conditions = "Pre-conditions"
        Test_data = "N/A"
        Test_Steps = "Open Site\nClick in New window\nValidar text " + self.texto
        Expected_Result = "Text should be in site"
        Tester = "tester" + str(filas)
        Actual_result = "Text isnt in front of the site"
        Status = "Fail"
        Comments = "Test automation failed"
        Test_Evidence = "Or the output"
        value_col = [Test_Case_ID, Developer, Test_group, Test_case_name, Priority_Severity, Pre_conditions, Test_data,
                     Test_Steps, Expected_Result, Tester, Actual_result, Status, Comments, Test_Evidence]
        for col in range(columnas):
            hoja.cell(row=int(filas + 1), column=col + 1, value=value_col[col])

            print("------------------------------------")
            print("The Excel File wa update: " + Config.Excel)
            print("------------------------------------")
        wb.save(Config.Excel)

    def take_screenshot(self, TestCase="Captura"):
        """
        Take a screenshot of the device

        :param
            -TestCase: name of testcase, extracted by the scenarioname
        """
        PATH = Config.basedir + "reports/screenshots"
        img = f'{PATH}/{TestCase}_(' + str(Functions.current_hours(self)) + ')' + '.png'
        self.driver.get_screenshot_as_file(img)

    def current_hours(self):
        self.hour = datetime.now().strftime("%H%M%S")
        return self.hour

    def scroll_to_element(self, obj):
        """
        This step helps to scrrolling to element selected

        :param
            obj: name of the element
        """
        try:
            element = Functions.get_elements(self, obj)
            self.driver.execute_script("arguments[0].scrollIntoView();", element)
            time.sleep(0.5)

        except:
            ActionChains(self.driver) \
                .scroll_to_element(element) \
                .perform()